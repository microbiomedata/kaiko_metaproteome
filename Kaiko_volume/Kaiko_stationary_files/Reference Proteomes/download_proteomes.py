import re
import os
import time
import requests
import pandas as pd
import datetime
 
from openpyxl import load_workbook
from pathlib import PureWindowsPath, Path

## Add flag for this
referece_proteomes_table_path = Path(PureWindowsPath('proteomes_AND_proteome_type_1_2023_11_28.tsv'))
proteome_table = pd.read_csv(referece_proteomes_table_path, sep = "\t")
proteome_table = proteome_table.set_index('Organism Id')

log_file = Path(PureWindowsPath('database_log.xlsx'))

## base pattern for accessing proteomes programatically
rest_uniprot_pattern_compressed = 'https://rest.uniprot.org/uniprotkb/stream?compressed=true&format=fasta&query=%28%28proteome%3A'
rest_uniprot_pattern_uncompressed = 'https://rest.uniprot.org/uniprotkb/stream?compressed=false&format=fasta&query=%28%28proteome%3A'


def download_proteome(taxa_id, rest_uniprot_pattern):
    proteome_id = proteome_table.loc[taxa_id]['Proteome Id']
    download_proteome_by_id(proteome_id, rest_uniprot_pattern, suffix = f'_taxaid_{taxa_id}')


def download_proteome_by_id(proteome_id, rest_uniprot_pattern, suffix = ''):
    rest_uniprot_url = f'{rest_uniprot_pattern}{proteome_id}%29%29'
    request = requests.get(rest_uniprot_url, stream = True)
    start_time = time.time() 
    with open(f'{proteome_id}{suffix}.fasta', 'wb') as f:
        # for chunk in request.raw.stream(1024, decode_content=True):
        for chunk in request.iter_content(chunk_size=2**20):
            if chunk:
                f.write(chunk)
        print(time.time()-start_time)
    

def log_proteomes(taxa_ids, dates, proteome_paths, log_file):
    if not log_file.exists():
        new_log = proteome_table[['Proteome Id', 'Organism', 'Taxonomic lineage', 'Genome assembly ID']].copy()
        new_log['Date downloaded'] = None
        new_log['FASTA size'] = None
        new_log.to_excel(str(log_file))
    log = load_workbook(log_file)

    sheet = log.active
    for index, taxa_id in enumerate(taxa_ids):
        proteome_path = proteome_paths[index]
        date = dates[index]
        taxa_index = proteome_table.index.to_list().index(taxa_id) 
        sheet.cell(row = taxa_index + 2, column = 6, value = date)
        if proteome_path.exists():
            fasta_size = os.path.getsize(proteome_path)
        else:
            fasta_size = 0
        if fasta_size == 0:
                print(f'Downloading {taxa_id} failed')
        sheet.cell(row = taxa_index + 2, column = 7, value = fasta_size)
    log.save(log_file)

taxa_ids = []
dates = []
proteome_paths = []
for taxa_id in proteome_table.index.to_list():
    proteome_id = proteome_table.loc[taxa_id]['Proteome Id']
    suffix = f'_taxaid_{taxa_id}'
    proteome_path = Path(PureWindowsPath(f'{proteome_id}{suffix}.fasta'))
    if not proteome_path.exists() or os.path.getsize(proteome_path) == 0:
        if not proteome_path.exists():
            print(f'Downloading taxa id {taxa_id}')
        else:
            print(f'Last attempt to download proteome of taxa id {taxa_id} failed. Retrying')
        try:
            download_proteome(taxa_id, rest_uniprot_pattern_uncompressed)
        except:
            print(f"Failed to download proteome of taxa {taxa_id}")
            log_proteomes(taxa_ids, dates, proteome_paths, log_file)
            os.remove(proteome_path)
            break
        taxa_ids = taxa_ids + [taxa_id]
        dates = dates + [datetime.datetime.now()]
        proteome_paths = proteome_paths + [proteome_path]
        if len(taxa_ids) >= 7:
            print(f"Logging last {len(taxa_ids)} downloads")
            log_proteomes(taxa_ids, dates, proteome_paths, log_file)
            taxa_ids = []
            dates = []
            proteome_paths = []
    else:
        print(f'Taxa {taxa_id} already downloaded')

print("Logging last downloads...")
log_proteomes(taxa_ids, dates, proteome_paths, log_file)
print("Finished downloading")